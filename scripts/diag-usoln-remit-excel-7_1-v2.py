"""Dump both 7월 1주차 excel variants — normal + (1) — to compare."""
import glob
import io
import os
import openpyxl

DOWNLOADS = os.path.expanduser("~/Downloads")
matches = sorted(glob.glob(os.path.join(DOWNLOADS, "유솔N_주정산_7월 1주차*.xlsx")))

out_path = os.path.join(os.path.dirname(__file__), "diag-usoln-remit-excel-7_1-v2.out")
buf = io.StringIO()

def log(*args, **kwargs):
    print(*args, **kwargs, file=buf)

log("=== 후보 파일 ===")
for m in matches:
    log(f"  {m}  size={os.path.getsize(m)}  mtime_epoch={os.path.getmtime(m)}")

for path in matches:
    log(f"\n===== {os.path.basename(path)} =====")
    log(f"경로: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log(f"\n--- 시트: {sheet_name}  rows={ws.max_row} cols={ws.max_column} ---")
        for r in range(1, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            log(f"  row{r}: {row}")

with open(out_path, "w", encoding="utf-8") as f:
    f.write(buf.getvalue())

print(f"Written: {out_path}")
