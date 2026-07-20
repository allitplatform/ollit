"""Generate backfill SQL files for 7/6 · 7/13 confirmed remits.

Reads two excel files:
  · 유솔N_주정산_7월 1주차 (1).xlsx  → 7/6 확정 (115건, 당주 only)
  · 유솔N_주정산_7월 2주차 (1).xlsx  → 7/13 확정 (53건, 당주 14 + 이월 39)

Emits:
  · db/migrations/ad-hoc_2026-07-20_backfill_remit_items_7_6.sql
  · db/migrations/ad-hoc_2026-07-20_backfill_remit_items_7_13.sql

Each SQL:
  1. Resolves remit_id from principals(code='usol_n') + week_start
  2. VALUES 절로 엑셀 items 인라인
  3. LEFT JOIN task_items ON product_order_id → 매칭
  4. INSERT INTO principal_weekly_remit_items (매칭된 것만)
  5. UPDATE snapshot_item_count = 엑셀 원본 건수
  6. UPDATE task_items.company_received_at (매칭된 것만, IS NULL 조건)
  7. RAISE NOTICE 로 matched/missing 카운트 리포트
  8. Transaction wrap (BEGIN ... COMMIT) — 문제 시 ROLLBACK

Snapshot 값 정책:
  · subtotal        = 엑셀 정산원금 (확정 시점 값 고정)
  · net_amount      = NULL (엑셀에 없음)
  · company_receive_amount = ROUND(subtotal × 0.85) — Mig 185 동일 산식
  · naver_settled_at = task_items.naver_settled_at (원본 timestamptz · 드릴인 정렬용)
  · captured_at     = 그 remit 의 confirmed_at (확정 시점 스냅샷 의미)
"""
import glob
import os
import re
import sys
import openpyxl

DOWNLOADS = os.path.expanduser("~/Downloads")
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
OUT_DIR = os.path.join(REPO_ROOT, "db", "migrations")

TARGETS = [
    {
        "label": "7_6",
        "pattern": "유솔N_주정산_7월 1주차 (1).xlsx",
        "week_start": "2026-06-29",
        "week_end":   "2026-07-05",
        "expected_count": 115,
        "expected_company_receive": 8_187_023,
        "expected_carryover_count": 0,
    },
    {
        "label": "7_13",
        "pattern": "유솔N_주정산_7월 2주차 (1).xlsx",
        "week_start": "2026-07-06",
        "week_end":   "2026-07-12",
        "expected_count": 53,
        "expected_company_receive": 2_865_597,
        "expected_carryover_count": 39,
    },
]

CARRYOVER_RE = re.compile(r"이전 미정산 \((\d+)/(\d+)~\d+/\d+\)")

def parse_row(row):
    """Return dict from 건별 상세 row (9 cols). settled_date not used — pulled from task_items."""
    kind = row[0]
    product_order_id = str(row[2]).strip() if row[2] is not None else None
    buyer_name = str(row[3]).strip() if row[3] is not None else ""
    subtotal = int(row[6]) if row[6] is not None else None

    if not product_order_id or subtotal is None:
        return None

    is_carryover = False
    carry_monday = None
    if kind != "당주 정산":
        m = CARRYOVER_RE.match(str(kind))
        if not m:
            raise ValueError(f"unrecognized 구분: {kind!r}")
        month, day = int(m.group(1)), int(m.group(2))
        carry_monday = f"2026-{month:02d}-{day:02d}"
        is_carryover = True

    return {
        "product_order_id": product_order_id,
        "buyer_name": buyer_name,
        "subtotal": subtotal,
        "is_carryover": is_carryover,
        "carry_monday": carry_monday,
    }

def load_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "건별 상세" not in wb.sheetnames:
        raise ValueError(f"sheet '건별 상세' not found in {path}. sheets={wb.sheetnames}")
    ws = wb["건별 상세"]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        parsed = parse_row(row)
        if parsed:
            rows.append(parsed)
    return rows

def format_values_row(item):
    poid = item["product_order_id"].replace("'", "''")
    buyer = item["buyer_name"].replace("'", "''")
    carry_sql = f"'{item['carry_monday']}'::date" if item["carry_monday"] else "NULL::date"
    return (
        f"    ('{poid}', '{buyer}', {item['subtotal']}, "
        f"{'true' if item['is_carryover'] else 'false'}, "
        f"{carry_sql})"
    )

def emit_sql(target, items):
    label = target["label"]
    week_start = target["week_start"]
    week_end = target["week_end"]
    expected_count = target["expected_count"]
    expected_recv = target["expected_company_receive"]
    expected_carry = target["expected_carryover_count"]

    actual_count = len(items)
    actual_carry = sum(1 for it in items if it["is_carryover"])
    actual_recv = sum(round(it["subtotal"] * 0.85) for it in items)

    values_lines = ",\n".join(format_values_row(it) for it in items)

    sql = f"""-- ad-hoc_2026-07-20_backfill_remit_items_{label}.sql
-- 2026-07-20 v3 — DO 블록 / RAISE NOTICE / TEMP TABLE / BEGIN·COMMIT 전부 제거
--
-- v2 실패 원인 (Supabase SQL Editor):
--   DO $$ 블록 실행 여부 확인 불가. "Success. No rows returned" 만 표시.
--   NOTICE 안 보임. 트랜잭션 컨텍스트 처리 방식 불명 → 실제 INSERT 0건.
--   진단 SQL [S3] 결과: 매칭·필터 모두 정상 (10_passes_ALL=53).
--
-- v3 구조: 순수 SQL 문장 4개.
--   [1] INSERT ... SELECT ... ON CONFLICT DO NOTHING RETURNING id
--       → 편집기에 rows returned 로 삽입 건수 표시
--   [2] UPDATE principal_weekly_remittances SET snapshot_item_count ... RETURNING
--       → 요약 세팅 확인
--   [3] UPDATE task_items SET company_received_at ... RETURNING id
--       → 스탬프 건수 표시
--   [4] SELECT 검증 — 최종 카운트 · 총액
--
-- 실행:
--   각 섹션 [1]~[4] 를 블록 지정 후 개별 Run.
--   각 문장 자동 커밋 (BEGIN/COMMIT 없음).
--   문제 발생 시 각 문장 결과 즉시 확인 가능.
--
-- 엑셀 파싱 결과:
--   · 총 건수: {actual_count} (기대: {expected_count})
--   · 이월 건수: {actual_carry} (기대: {expected_carry})
--   · Σ ROUND(subtotal × 0.85): {actual_recv:,} (기대: {expected_recv:,})

-- ============================================================================
-- [1] 스냅샷 INSERT
--   기대: {actual_count} rows returned (편집기 하단 rows count 확인)
--   0 rows returned → remit 조회 실패 or 필터 이상 (진단 SQL [S3] 재확인)
--   {actual_count} 미만 → 매칭 실패분 (진단 SQL [S4] 로 상세 확인)
-- ============================================================================
WITH remit AS (
  SELECT id AS remit_id, confirmed_at, tenant_id
    FROM principal_weekly_remittances
   WHERE principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
     AND week_start = DATE '{week_start}'
     AND confirmed_at IS NOT NULL
),
excel(poid, buyer_name, subtotal, is_carryover, carry_monday) AS (
  VALUES
{values_lines}
)
INSERT INTO principal_weekly_remit_items (
  tenant_id, remit_id, task_item_id,
  subtotal, net_amount, company_receive_amount,
  is_carryover, carryover_source_monday,
  naver_settled_at, captured_at
)
SELECT
  r.tenant_id,
  r.remit_id,
  ti.id,
  e.subtotal,
  NULL,
  ROUND(e.subtotal::numeric * 0.85)::int,
  e.is_carryover,
  e.carry_monday,
  ti.naver_settled_at,
  r.confirmed_at
FROM excel e
JOIN task_items ti ON ti.product_order_id = e.poid
JOIN tasks       t ON t.id = ti.task_id
CROSS JOIN remit r
WHERE t.principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
  AND (t.status IS NULL OR t.status != '취소')
  AND COALESCE(ti.is_canceled, false) = false
ON CONFLICT (remit_id, task_item_id) DO NOTHING
RETURNING id;

-- ============================================================================
-- [2] snapshot_item_count 세팅 (엑셀 원본 건수 · 매칭 실패해도 사장님 값 유지)
--   기대: 1 row returned (그 remit row 정보 표시)
-- ============================================================================
UPDATE principal_weekly_remittances
   SET snapshot_item_count = {actual_count},
       updated_at = now()
 WHERE principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
   AND week_start = DATE '{week_start}'
   AND confirmed_at IS NOT NULL
RETURNING id, week_start, confirmed_at, remitted_amount, snapshot_item_count;

-- ============================================================================
-- [3] company_received_at 스탬프 (스냅샷 items 중 NULL 만)
--   기대 ({label}):
--     · 7_13 → 이월 39건 스탬프 (당주 14 는 이미 다른 경로 스탬프됨)
--     · 7_6  → 4건 or 이하 (사장님 언급 111건 이미 스탬프 · 나머지 4건 유실 처리)
--   실제 rows returned 로 확인.
-- ============================================================================
UPDATE task_items ti
   SET company_received_at = (
     SELECT confirmed_at FROM principal_weekly_remittances
      WHERE principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
        AND week_start = DATE '{week_start}'
        AND confirmed_at IS NOT NULL
   )
 WHERE ti.id IN (
   SELECT ri.task_item_id
     FROM principal_weekly_remit_items ri
    WHERE ri.remit_id = (
      SELECT id FROM principal_weekly_remittances
       WHERE principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
         AND week_start = DATE '{week_start}'
         AND confirmed_at IS NOT NULL
    )
 )
   AND ti.company_received_at IS NULL
RETURNING ti.id;

-- ============================================================================
-- [4] 최종 검증 — 스냅샷 카운트 · 총액 · 이월 분리
--   기대: cnt={actual_count} · sum_recv≈{actual_recv:,} · carry_cnt={actual_carry}
--   (반올림 차이 ±수원 정상)
-- ============================================================================
SELECT
  count(*)                                      AS cnt,
  sum(company_receive_amount)                   AS sum_recv,
  sum(CASE WHEN is_carryover THEN 1 ELSE 0 END) AS carry_cnt,
  sum(subtotal)                                 AS sum_subtotal
FROM principal_weekly_remit_items
WHERE remit_id = (
  SELECT id FROM principal_weekly_remittances
   WHERE principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
     AND week_start = DATE '{week_start}'
);

-- ============================================================================
-- [5] (선택) 이월 소멸 확인 — 소급 후 미회수 items 재계산
--   기대: {label} 소급 완료 후 그 주차까지의 미회수 이월 감소 확인
-- ============================================================================
-- SELECT count(*) AS remaining_unpaid_upto_{label},
--        sum(ti.subtotal) AS remaining_sum
--   FROM task_items ti
--   JOIN tasks t ON ti.task_id = t.id
--  WHERE t.principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
--    AND t.status = '완료'
--    AND ti.naver_settled_at IS NOT NULL
--    AND ti.company_received_at IS NULL
--    AND (ti.naver_settled_at AT TIME ZONE 'Asia/Seoul')::date <= DATE '{week_end}'
--    AND COALESCE(ti.is_canceled, false) = false
--    AND ti.subtotal > 0
--    AND EXISTS (SELECT 1 FROM payments p WHERE p.task_id = t.id AND p.track = 'B');
"""
    out_path = os.path.join(OUT_DIR, f"ad-hoc_2026-07-20_backfill_remit_items_{label}.sql")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(sql)
    return out_path, actual_count, actual_carry, actual_recv

def emit_diag_sql(target, items):
    """Emit SELECT-only diagnostic SQL — no BEGIN/COMMIT, no TEMP tables, no DO block.
    Supabase SQL Editor 는 마지막 SELECT 결과만 표시 → 섹션별 개별 실행."""
    label = target["label"]
    week_start = target["week_start"]
    week_end = target["week_end"]

    def val_row(item):
        poid = item["product_order_id"].replace("'", "''")
        return f"    ('{poid}'::text, {item['subtotal']}::int)"

    values_lines = ",\n".join(val_row(it) for it in items)

    sql = f"""-- ad-hoc_2026-07-20_diag_backfill_{label}.sql
-- 2026-07-20 — 진단 SQL (SELECT-only, BEGIN/COMMIT/TEMP TABLE 없음)
--
-- Supabase SQL Editor 는 마지막 SELECT 만 결과 표시 → 아래 섹션 [S1]~[S5] 를
-- 각각 개별 실행. 각 섹션 끝의 세미콜론 앞까지 블록 지정 후 Run.
-- 결과 테이블 형태로 반환 → NOTICE 안 보이는 문제 회피.

-- ============================================================================
-- [S1] 컬럼 타입 확인 (product_order_id 실제 타입 · 매칭 실패 원인 후보)
-- ============================================================================
SELECT
  attname   AS column_name,
  format_type(atttypid, atttypmod) AS column_type
  FROM pg_attribute
 WHERE attrelid = 'task_items'::regclass
   AND attname IN ('product_order_id', 'company_received_at', 'is_canceled', 'subtotal')
   AND attnum > 0
 ORDER BY attnum;

-- ============================================================================
-- [S2] remit row 정보 (id · week_start · confirmed_at · remitted_amount)
-- ============================================================================
SELECT id, week_start, week_end, confirmed_at, remitted_amount,
       snapshot_item_count
  FROM principal_weekly_remittances
 WHERE principal_id = (SELECT id FROM principals WHERE code = 'usol_n')
   AND week_start = DATE '{week_start}';

-- ============================================================================
-- [S3] 매칭 단계별 카운트 요약 (핵심 진단)
--   기대 (7_13): total=53, ti_matched=53, task_matched=53, passes_all=53
--   실제와 대조 → 어느 단계에서 유실됐는지 즉시 파악
-- ============================================================================
WITH usoln AS (SELECT id AS pid FROM principals WHERE code = 'usol_n'),
     excel(poid, subtotal) AS (
       VALUES
{values_lines}
     ),
     joined AS (
       SELECT
         e.poid                                     AS excel_poid,
         e.subtotal                                 AS excel_subtotal,
         ti.id                                      AS ti_id,
         ti.product_order_id                        AS ti_poid,
         ti.company_received_at                     AS ti_stamped,
         ti.is_canceled                             AS ti_canceled,
         t.id                                       AS task_id,
         t.principal_id                             AS task_principal,
         t.status                                   AS task_status
       FROM excel e
       LEFT JOIN task_items ti ON ti.product_order_id = e.poid
       LEFT JOIN tasks       t ON t.id = ti.task_id
     ),
     existing AS (
       SELECT task_item_id FROM principal_weekly_remit_items
        WHERE remit_id = (
          SELECT id FROM principal_weekly_remittances
           WHERE principal_id = (SELECT pid FROM usoln)
             AND week_start = DATE '{week_start}'
        )
     )
SELECT '01_excel_total'                            AS bucket, count(*) AS cnt FROM joined
UNION ALL SELECT '02_ti_matched',                    count(*) FROM joined WHERE ti_id IS NOT NULL
UNION ALL SELECT '03_ti_unmatched (poid mismatch)',  count(*) FROM joined WHERE ti_id IS NULL
UNION ALL SELECT '04_task_matched',                  count(*) FROM joined WHERE task_id IS NOT NULL
UNION ALL SELECT '05_task_orphan (ti O + task X)',   count(*) FROM joined WHERE ti_id IS NOT NULL AND task_id IS NULL
UNION ALL SELECT '06_principal_is_usoln',            count(*) FROM joined WHERE task_principal = (SELECT pid FROM usoln)
UNION ALL SELECT '07_principal_other',               count(*) FROM joined WHERE task_principal IS NOT NULL AND task_principal != (SELECT pid FROM usoln)
UNION ALL SELECT '08_status_not_cancelled',          count(*) FROM joined WHERE task_status != '취소' OR task_status IS NULL
UNION ALL SELECT '09_ti_not_cancelled',              count(*) FROM joined WHERE COALESCE(ti_canceled, false) = false
UNION ALL SELECT '10_passes_ALL_filters',            count(*) FROM joined
  WHERE ti_id IS NOT NULL
    AND task_principal = (SELECT pid FROM usoln)
    AND (task_status != '취소' OR task_status IS NULL)
    AND COALESCE(ti_canceled, false) = false
UNION ALL SELECT '11_already_snapshot (would conflict)', count(*) FROM joined
  WHERE ti_id IN (SELECT task_item_id FROM existing)
ORDER BY 1;

-- ============================================================================
-- [S4] 실패 샘플 상세 (앞 20건)
--   ti_matched / task_matched / task_principal (usol_n 여부) / task_status /
--   ti_stamped (company_received_at) 컬럼으로 실패 사유 판단.
-- ============================================================================
WITH usoln AS (SELECT id AS pid FROM principals WHERE code = 'usol_n'),
     excel(poid, subtotal) AS (
       VALUES
{values_lines}
     )
SELECT
  e.poid                                           AS excel_poid,
  e.subtotal                                       AS excel_subtotal,
  ti.id IS NOT NULL                                AS ti_matched,
  t.id IS NOT NULL                                 AS task_matched,
  (t.principal_id = (SELECT pid FROM usoln))       AS is_usoln,
  t.status                                         AS task_status,
  ti.is_canceled                                   AS ti_canceled,
  ti.company_received_at                           AS ti_stamped
FROM excel e
LEFT JOIN task_items ti ON ti.product_order_id = e.poid
LEFT JOIN tasks       t ON t.id = ti.task_id
WHERE ti.id IS NULL
   OR t.principal_id IS NULL
   OR t.principal_id != (SELECT pid FROM usoln)
   OR t.status = '취소'
   OR COALESCE(ti.is_canceled, false)
LIMIT 20;

-- ============================================================================
-- [S5] 성공 샘플 상세 (앞 5건 — 실제로 매칭 · 필터 통과했는지 눈으로 확인)
-- ============================================================================
WITH usoln AS (SELECT id AS pid FROM principals WHERE code = 'usol_n'),
     excel(poid, subtotal) AS (
       VALUES
{values_lines}
     )
SELECT
  e.poid          AS excel_poid,
  e.subtotal      AS excel_subtotal,
  ti.id           AS ti_id,
  ti.subtotal     AS ti_subtotal,
  t.customer_name AS customer_name,
  t.status        AS task_status,
  ROUND(e.subtotal::numeric * 0.85)::int AS expected_company_receive
FROM excel e
JOIN task_items ti ON ti.product_order_id = e.poid
JOIN tasks       t ON t.id = ti.task_id
WHERE t.principal_id = (SELECT pid FROM usoln)
  AND t.status != '취소'
  AND COALESCE(ti.is_canceled, false) = false
LIMIT 5;
"""
    out_path = os.path.join(OUT_DIR, f"ad-hoc_2026-07-20_diag_backfill_{label}.sql")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(sql)
    return out_path

def main():
    for target in TARGETS:
        matches = glob.glob(os.path.join(DOWNLOADS, target["pattern"]))
        if not matches:
            print(f"[skip] {target['label']}: file not found ({target['pattern']})", file=sys.stderr)
            continue
        path = matches[0]
        items = load_excel(path)
        out_path, cnt, carry, recv = emit_sql(target, items)
        exp_cnt = target["expected_count"]
        exp_carry = target["expected_carryover_count"]
        exp_recv = target["expected_company_receive"]
        status = "OK" if (cnt == exp_cnt and carry == exp_carry) else "MISMATCH"
        recv_status = "OK" if recv == exp_recv else f"DIFF({recv - exp_recv:+d})"
        print(f"[{status}] {target['label']}: count={cnt}/{exp_cnt} carry={carry}/{exp_carry} recv={recv:,}/{exp_recv:,} {recv_status}")
        print(f"        emitted backfill: {out_path}")
        diag_path = emit_diag_sql(target, items)
        print(f"        emitted diag:     {diag_path}")

if __name__ == "__main__":
    main()
